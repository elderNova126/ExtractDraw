from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import json
import zipfile
import re

def clean_tolerance_string(tol):
    # Remove tolerance markers like H7, T3, J2, etc. (letter followed by digits)
    tol = re.sub(r'[A-Za-z]\d+', '', tol)
    
    # Remove symbols like ▽ followed by a number (e.g., ▽5, ▽10)
    tol = re.sub(r'[^\w\s\+\-\/\.\±]+[\d]+', '', tol)
    
    return tol.strip()
def parse_tolerance(dim, tol):
    if not dim:  # If dim is None or empty
        return "", "", ""

    try:
        dim = dim.strip().replace(" ", "")
    except:
        dim = ""
    try:
        tol = tol.strip().replace(" ", "")
    except:
        tol = ""

    dim = dim.strip().replace(" ", "")
    
    # Remove unwanted characters (e.g., ▽4)
    tol = clean_tolerance_string(tol)

    if dim == '⌀' and tol:
        return dim, tol, tol

    match = re.findall(r'(\d+\.?\d*)', dim)
    if not match:
        return dim, "", ""
    
    if "x" in dim and len(match) > 1:
        base_dim = float(match[1])  # Second number should be the reference dimension
    else:
        base_dim = float(match[0])

    dim = f"⌀{base_dim}" if '⌀' in dim else str(base_dim)
    
    if not tol:
        return dim, f"{base_dim:.3f}", f"{base_dim:.3f}"
    
    match = re.search(r'^([A-Za-z0-9]*)\s*([-+]?\d*\.?\d+/[-+]?\d*\.?\d+|[-+]?\d*\.?\d*|±\d*\.?\d+)$', tol)
    if match:
        prefix, tol_values = match.groups()
    else:
        return dim, f"{base_dim:.3f}", f"{base_dim:.3f}"
    
    tol_values = re.findall(r'[-+]?\d*\.?\d+', tol_values)
    
    if len(tol_values) == 2:  # Asymmetric tolerance (e.g., "0/-0.09")
        try:
            upper_tol, lower_tol = max(float(tol_values[0]), float(tol_values[1])), min(float(tol_values[0]), float(tol_values[1]))
            return dim, f"{base_dim + upper_tol:.3f}", f"{base_dim + lower_tol:.3f}"
        except ValueError:
            return dim, "", ""
    
    if len(tol_values) == 1:  # Single tolerance value case (e.g., "±0.05")
        try:
            single_tol = float(tol_values[0])
            if "±" in tol:
                return dim, f"{base_dim + single_tol:.3f}", f"{base_dim - single_tol:.3f}"
            return dim, f"{max(base_dim + single_tol, base_dim):.3f}", f"{min(base_dim + single_tol, base_dim):.3f}"
        except ValueError:
            return dim, "", ""
    
    return dim, f"{base_dim:.3f}", f"{base_dim:.3f}"

def post_processing(All_results, savePath, zip_filename):
    '''
    This function makes xlsx files from Boxes and Texts.
    1. Loop Boxes and Texts according to document.
    2. In a document, loop boxes and texts according to page.
    3. SR number modification
    4. Consider cell swrap, thin
    5. Save
    '''

    for ind, results in enumerate(All_results):
        print("<========> ", results)
        dimension_string, [drawing_num, drawing_scale] = results
        try:
            dimensions = json.loads(dimension_string)  
            col_title = ['項目', '規格', '公差', '上限', '下限', '工具']
            pre_rows = 8 # considering multi tables in a page.
            wb = Workbook()
            ws = wb.active
            ws.title = "new table"
            # for i in range(len(col_title)):
            #     ws.cell(pre_rows+1,i+1).value = col_title[i]
            #     ws.cell(pre_rows+1,i+1).font = Font(bold=True)    
            thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')) 

            ws.merge_cells('A1:M1')
            ws['A1'] = "机械有限公司"
            ws.merge_cells('A2:M2')
            ws['A2'] = "出貨檢驗報告"
            # Third row
            ws.merge_cells('A3:B3') 
            ws['A3'] = "客户"
            ws.merge_cells('C3:F3')
            ws['C3'] = ""
            ws.merge_cells('G3:H3')  
            ws['G3'] = "A39_輸送   SL_全氣式"
            ws.merge_cells('I3:M3')
            ws['I3'] = ""

            # Fourth row
            ws.merge_cells('A4:B4') 
            ws['A4'] = "訂單編號:"
            ws.merge_cells('C4:F4')  
            ws['C4'] = ""
            ws.merge_cells('G4:H4')
            ws['G4'] = "圖號："
            ws.merge_cells('I4:K4')  
            ws['I4'] = drawing_num
            ws.merge_cells('L4:M4') 
            ws['L4'] = drawing_scale
            # Fifth row
            ws.merge_cells('A5:B5') 
            ws['A5'] = "出貨數量："
            ws.merge_cells('C5:F5')
            ws['C5'] = ""
            ws.merge_cells('G5:H5')  
            ws['G5'] = "檢驗日期："
            ws.merge_cells('I5:M5')
            ws['I5'] = ""

            # Sixth row
            ws.merge_cells('A6:B6') 
            ws['A6'] = "尺寸抽樣水準："
            ws.merge_cells('C6:F6')
            ws['C6'] = "依抽樣計畫"
            ws.merge_cells('G6:H6')  
            ws['G6'] = "外觀抽樣水準："
            ws.merge_cells('I6:M6')
            ws['I6'] = "MIL-STD-1916  VL Ⅱ"

            # Seventh row
            ws.merge_cells('A7:B7') 
            ws['A7'] = "尺寸抽樣數："
            ws['C7'] = ""
            ws['D7'] = "PCS"
            ws.merge_cells('E7:F7') 
            ws['E7'] = ""
            ws.merge_cells('G7:H7')  
            ws['G7'] = "外觀抽樣數："
            ws.merge_cells('I7:J7')
            ws['I7'] = ""
            ws['K7'] = ""
            ws.merge_cells('L7:M7')
            ws['L7'] = ""

            # Eighth row
            ws['A8'] = "項目"
            ws.merge_cells('B8:C8')  
            ws['B8'] = "規格/公差"
            ws.merge_cells('D8:E8')  
            ws['D8'] = "上限/下限"
            ws['F8'] = "工具"
            ws.merge_cells('G8:K8')  
            ws['G8'] = "實測值"
            ws['L8'] = "OK"
            ws['M8'] = "NotOK"            
            for key, value in dimensions.items():
                try:
                    ws.cell(row=pre_rows+1, column=1).value = key
                    ws.cell(row=pre_rows+1, column=3).value = value["tolerance"]
                    dim_value, up_limit, low_limit = parse_tolerance(value['dim'], value['tolerance'])
                    ws.cell(row=pre_rows+1, column=2).value = dim_value
                    ws.cell(row=pre_rows+1, column=4).value = up_limit
                    ws.cell(row=pre_rows+1, column=5).value = low_limit        
                    ws.cell(row=pre_rows+1, column=6).value = "三次元"
                    # ws.cell(row=pre_rows+1, column=7).value = f'{value["dim"]} {value["tolerance"]}'
                    pre_rows += 1   
                except: pass
            # cell swrap, thin
            row_no = 1
            for i in ws.rows:
                for j in range(len(i)):
                    ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
                    ws.cell(row=row_no, column=j + 1).border = thin_border
                row_no = row_no + 1  
            column_width = 15*[13]

            for i in range(len(col_title)):
                ws.column_dimensions[get_column_letter(i+1)].width = column_width[i]
            ws.sheet_view.zoomScale = 85
            print("======> ", ind)
            wb.save(savePath[ind])                
        except Exception as e:
            print(f"Error in post processing : {e}")
            pass
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in savePath:
            try:
                zipf.write(file, arcname=file.split('\\')[-1])
            except: pass

    return None

# sample_string = ('{    "1": {"dim": "3", "tolerance": "+0.008/+0.002 ▽4"},    "2": {"dim": "82.5", "tolerance": ""},    "3": {"dim": "3", "tolerance": "+0.008/+0.002"},    "4": {"dim": "⌀4", "tolerance": "H7 +0.012/0 ▽5"},    "5": {"dim": "39", "tolerance": ""},    "6": {"dim": "117.6", "tolerance": ""},    "7": {"dim": "31.4", "tolerance": "+0.05"},    "8": {"dim": "199", "tolerance": ""},    "9": {"dim": "10", "tolerance": ""},    "10": {"dim": "54", "tolerance": "+0.1/0"}}', ['819101000563', '1.00'])
# post_processing([sample_string], ["a.xlsx"], ["a.zip"])
